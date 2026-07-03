"""
pay_app.py  —  Payment console for the Teen AI Survey.
==============================================================================
A tiny local web app (Python stdlib + openpyxl, both already in the shared
venv) that handles the *human* part of paying completed participants:

  1. Shows everyone who is payable and not yet paid — the same "Pay (no hold)"
     population that manage_payments.py already computed (fraud + holds excluded).
  2. Lets you pick a batch (all of them, or a subset you define by pasting
     emails) and COPY their addresses to the clipboard, so you can paste them
     straight into Amazon's "one email per recipient" gift-card box (it accepts
     up to 999 addresses separated by a comma or a space).
  3. After you buy + send the $10 cards on Amazon, you click "Mark paid + email"
     on that batch. The app:
       - writes  paid=yes, paid_date, paid_amount  back into
         data/payment_tracker.xlsx  (sheet "Payments", keyed by cid) — exactly
         the manual edit the README asks you to make, so the NEXT
         manage_payments.py run drops these people from the report.
       - sends each person the thank-you email through classic Outlook (the
         same shared airlab@uchicago.edu mailbox the rest of the pipeline uses),
         personalized with their first name.
       - logs the payment + email to  data/payment_email_log.csv  (app-owned).
  4. A bookkeeping table lists everyone paid through the app and lets you tick
     "received confirmed" when they reply "I've received my $10 gift card".

It reimplements no payment/quality logic. The payable list comes from the
report manage_payments.py already writes; this app just drives the manual step
and records the result.

Run with run_pay_app.bat (double-click), or:
    ..\\.venv\\Scripts\\python.exe pay_app.py
Then open http://127.0.0.1:5001 (opens automatically).

Requirements: Windows + classic Outlook (NOT New Outlook) open and signed in as
your UChicago account, for the email step. Everything else works without it.
==============================================================================
"""

import csv
import json
import threading
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

try:
    import yaml
except Exception:
    yaml = None

try:
    import openpyxl
except Exception:
    openpyxl = None

# -- Paths --------------------------------------------------------------------
HERE = Path(__file__).resolve().parent          # survey_management/payment_management
BASE = HERE.parent                              # survey_management
CONFIG_PATH = BASE / "config.yaml"
TRACKER_PATH = BASE / "data" / "payment_tracker.xlsx"          # ledger (manage_payments.py owns it)
REPORT_PATH = BASE / "data" / "payment_report_unpaid.xlsx"     # Pay/Hold/Fraud report
EMAIL_LOG_PATH = BASE / "data" / "payment_email_log.csv"       # this app owns it
TEMPLATE_PATH = HERE / "pay_app.html"

PORT = 5001

# A "paid" cell counts as paid if it holds any of these (mirrors dashboard/app.py).
PAID_TRUE_VALUES = {"yes", "y", "true", "1", "paid", "done", "sent", "complete"}

PAY_SHEET = "Pay (no hold)"
LEDGER_SHEET = "Payments"

# Columns the app writes/reads in the ledger (must exist in the Payments header).
LEDGER_KEY = "cid"
LEDGER_PAID = "paid"
LEDGER_PAID_DATE = "paid_date"
LEDGER_PAID_AMOUNT = "paid_amount"
LEDGER_NOTES = "notes"

EMAIL_LOG_FIELDS = [
    "cid", "first_name", "delivery_email", "paid_date", "paid_amount",
    "email_status", "emailed_at", "received_confirmed", "received_confirmed_at", "error",
]

# Serialize all spreadsheet reads/writes so a POST that is rewriting the xlsx
# never races a GET that is reading it.
IO_LOCK = threading.Lock()


# -- Config -------------------------------------------------------------------
def load_config():
    cfg = {}
    if yaml is not None and CONFIG_PATH.exists():
        try:
            cfg = yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")) or {}
        except Exception:
            cfg = {}
    sender = cfg.get("sender", {}) or {}
    return {
        "shared_mailbox": cfg.get("shared_mailbox") or "airlab@uchicago.edu",
        "contact_email": sender.get("contact_email") or "airlab@uchicago.edu",
        "sender_name": sender.get("name") or "The AirLab Team",
    }


CONFIG = load_config()


# -- Email template -----------------------------------------------------------
def build_email(first_name, amount, contact):
    """Return (subject, html_body) for the thank-you email."""
    amt = _fmt_amount(amount)
    name = (first_name or "there").strip() or "there"
    subject = f"Your ${amt} Amazon gift card - AirLab Teen AI Survey"
    body = f"""<html><body style="font-family:Calibri,Arial,sans-serif;font-size:11pt;color:#222;line-height:1.4;">
<p>Dear {_esc(name)},</p>
<p>Thank you for taking part in our survey. We really appreciate your time. As a
small token of our thanks, we've sent a ${amt} Amazon gift card to this email
address. It comes from Amazon, so if you don't see it shortly, please check your
spam or promotions folder.</p>
<p>To help us keep our records straight, please reply to this email with
<b>&quot;I've received my ${amt} gift card&quot;</b> once it arrives. If the card
doesn't show up or anything looks off, just reply here and we'll fix it.</p>
<p>Thanks again,<br>
The AirLab Team<br>
University of Chicago<br>
<a href="mailto:{_esc(contact)}">{_esc(contact)}</a></p>
</body></html>"""
    return subject, body


def _fmt_amount(amount):
    try:
        f = float(amount)
        return str(int(f)) if f == int(f) else f"{f:.2f}"
    except Exception:
        return str(amount)


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


# -- Outlook send (mirrors consent_management/send_survey_emails.py) ----------
_OUTLOOK = {}


def _get_outlook():
    import win32com.client as win32
    if "app" not in _OUTLOOK:
        _OUTLOOK["app"] = win32.Dispatch("outlook.application")
    return _OUTLOOK["app"]


def _shared_drafts_folder():
    mbx = CONFIG["shared_mailbox"]
    if not mbx:
        return None
    if "drafts" in _OUTLOOK:
        return _OUTLOOK["drafts"]
    ns = _get_outlook().GetNamespace("MAPI")
    recip = ns.CreateRecipient(mbx)
    recip.Resolve()
    if not recip.Resolved:
        raise RuntimeError(
            f"Outlook could not resolve shared mailbox '{mbx}'. Check the address "
            "and that the mailbox is added to your Outlook profile.")
    OL_FOLDER_DRAFTS = 16
    folder = ns.GetSharedDefaultFolder(recip, OL_FOLDER_DRAFTS)
    _OUTLOOK["drafts"] = folder
    return folder


def send_via_outlook(to_email, subject, html_body, draft_only):
    """Send (or draft) one message. Raises on any Outlook problem."""
    mail = _get_outlook().CreateItem(0)  # olMailItem
    mail.To = to_email
    mail.Subject = subject
    mail.HTMLBody = html_body
    mbx = CONFIG["shared_mailbox"]
    if mbx:
        mail.SentOnBehalfOfName = mbx
    if draft_only:
        mail.Save()
        folder = _shared_drafts_folder()
        if folder is not None:
            mail.Move(folder)
    else:
        mail.Send()


# -- Spreadsheet helpers ------------------------------------------------------
def _assert_writable(path: Path, label: str):
    """Raise a clear error if `path` is locked (open in Excel). Writes nothing."""
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        if path.exists():
            with open(path, "a"):
                pass
        else:
            probe = path.with_name(path.name + ".writetest")
            with open(probe, "w"):
                pass
            probe.unlink()
    except Exception as e:
        raise RuntimeError(
            f"Cannot write {label} ({path.name}). It is most likely open in Excel "
            f"- close it and try again. ({e})")


def _header_map(ws):
    """Return {header_text: column_index (1-based)} for row 1."""
    m = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            m[str(cell.value).strip()] = j
    return m


def _truthy_paid(v):
    return v is not None and str(v).strip().lower() in PAID_TRUE_VALUES


def read_ledger():
    """Return {cid: {first_name, child_name, delivery_email, paid(bool), paid_date,
    paid_amount, notes}} from the Payments sheet."""
    out = {}
    if openpyxl is None or not TRACKER_PATH.exists():
        return out
    wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True, data_only=True)
    if LEDGER_SHEET not in wb.sheetnames:
        wb.close()
        return out
    ws = wb[LEDGER_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return out
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def g(row, key):
        i = idx.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    for row in rows:
        cid = g(row, LEDGER_KEY)
        if not cid:
            continue
        out[str(cid).strip()] = {
            "first_name": g(row, "first_name"),
            "child_name": g(row, "child_name"),
            "delivery_email": g(row, "delivery_email"),
            "paid": _truthy_paid(g(row, LEDGER_PAID)),
            "paid_date": g(row, LEDGER_PAID_DATE),
            "paid_amount": g(row, LEDGER_PAID_AMOUNT),
            "notes": g(row, LEDGER_NOTES),
        }
    wb.close()
    return out


def read_pay_sheet():
    """Return list of candidate dicts from the 'Pay (no hold)' report sheet."""
    out = []
    if openpyxl is None or not REPORT_PATH.exists():
        return out
    wb = openpyxl.load_workbook(REPORT_PATH, read_only=True, data_only=True)
    if PAY_SHEET not in wb.sheetnames:
        wb.close()
        return out
    ws = wb[PAY_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return out
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def g(row, key):
        i = idx.get(key)
        return row[i] if (i is not None and i < len(row)) else None

    for row in rows:
        cid = g(row, "cid")
        if not cid:
            continue
        out.append({
            "cid": str(cid).strip(),
            "first_name": g(row, "first_name"),
            "delivery_email": (g(row, "delivery_email") or "").strip(),
            "n_flags": g(row, "n_flags"),
            "flag_reasons": g(row, "flag_reasons"),
            "survey_completed_at": str(g(row, "survey_completed_at") or ""),
        })
    wb.close()
    return out


def write_paid(cids, amount):
    """Set paid=yes / paid_date / paid_amount for each cid in the Payments sheet.
    Preserves every other cell. Returns (updated_cids, missing_cids)."""
    _assert_writable(TRACKER_PATH, "the payment tracker")
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb[LEDGER_SHEET]
    hm = _header_map(ws)
    for req in (LEDGER_KEY, LEDGER_PAID, LEDGER_PAID_DATE, LEDGER_PAID_AMOUNT):
        if req not in hm:
            wb.close()
            raise RuntimeError(f"Column '{req}' not found in the Payments sheet header.")
    key_c = hm[LEDGER_KEY]
    today = datetime.now().strftime("%Y-%m-%d")
    want = {str(c).strip() for c in cids}
    updated, seen = [], set()
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=key_c).value
        cid = str(cell).strip() if cell is not None else ""
        if cid in want:
            seen.add(cid)
            ws.cell(row=r, column=hm[LEDGER_PAID]).value = "yes"
            dcell = ws.cell(row=r, column=hm[LEDGER_PAID_DATE])
            if not dcell.value:
                dcell.value = today
            ws.cell(row=r, column=hm[LEDGER_PAID_AMOUNT]).value = float(amount)
            if LEDGER_NOTES in hm:
                ncell = ws.cell(row=r, column=hm[LEDGER_NOTES])
                stamp = f"paid via pay_app {today}"
                ncell.value = stamp if not ncell.value else f"{ncell.value}; {stamp}"
            updated.append(cid)
    wb.save(TRACKER_PATH)
    wb.close()
    missing = sorted(want - seen)
    return updated, missing


# -- Email-log CSV (app-owned; keyed by cid, rewritten in place) --------------
def read_email_log():
    rows = {}
    if not EMAIL_LOG_PATH.exists():
        return rows
    with open(EMAIL_LOG_PATH, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cid = (row.get("cid") or "").strip()
            if cid:
                rows[cid] = row
    return rows


def write_email_log(rows_by_cid):
    _assert_writable(EMAIL_LOG_PATH, "payment_email_log.csv")
    with open(EMAIL_LOG_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=EMAIL_LOG_FIELDS)
        w.writeheader()
        for cid in rows_by_cid:
            row = rows_by_cid[cid]
            w.writerow({k: row.get(k, "") for k in EMAIL_LOG_FIELDS})


# -- State assembly -----------------------------------------------------------
def build_state():
    with IO_LOCK:
        ledger = read_ledger()
        candidates = read_pay_sheet()
        log = read_email_log()

    # Payable = on the Pay sheet, not already paid in the ledger.
    payable = []
    for c in candidates:
        led = ledger.get(c["cid"])
        if led and led["paid"]:
            continue
        if not c["delivery_email"] and led:
            c["delivery_email"] = (led.get("delivery_email") or "").strip()
        if not c["first_name"] and led:
            c["first_name"] = led.get("first_name")
        payable.append(c)

    # Flag duplicate delivery emails within the payable set (two participants,
    # one inbox - usually a parent covering two kids; both should still be paid).
    seen_emails = {}
    for c in payable:
        e = c["delivery_email"].lower()
        seen_emails[e] = seen_emails.get(e, 0) + 1
    for c in payable:
        c["dup_email"] = seen_emails.get(c["delivery_email"].lower(), 0) > 1
        c["no_email"] = not c["delivery_email"]

    # Paid list = everyone the app has recorded in the email log.
    paid = []
    for cid, row in log.items():
        led = ledger.get(cid, {})
        paid.append({
            "cid": cid,
            "first_name": row.get("first_name") or led.get("first_name"),
            "delivery_email": row.get("delivery_email") or led.get("delivery_email"),
            "paid_date": row.get("paid_date"),
            "paid_amount": row.get("paid_amount"),
            "email_status": row.get("email_status"),
            "emailed_at": row.get("emailed_at"),
            "received_confirmed": (row.get("received_confirmed") or "").strip().lower() in ("yes", "true", "1"),
            "received_confirmed_at": row.get("received_confirmed_at"),
            "error": row.get("error"),
        })
    paid.sort(key=lambda r: (r["received_confirmed"], (r["first_name"] or "").lower()))

    counts = {
        "payable": len(payable),
        "paid_via_app": len(paid),
        "emailed_ok": sum(1 for p in paid if p["email_status"] in ("sent", "draft")),
        "email_failed": sum(1 for p in paid if p["email_status"] == "failed"),
        "confirmed": sum(1 for p in paid if p["received_confirmed"]),
    }
    return {
        "counts": counts,
        "payable": payable,
        "paid": paid,
        "config": {
            "shared_mailbox": CONFIG["shared_mailbox"],
            "contact_email": CONFIG["contact_email"],
            "default_amount": 10,
        },
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "tracker_found": TRACKER_PATH.exists(),
        "report_found": REPORT_PATH.exists(),
    }


# -- Actions ------------------------------------------------------------------
def action_mark_paid(payload):
    cids = [str(c).strip() for c in payload.get("cids", []) if str(c).strip()]
    if not cids:
        return {"ok": False, "error": "No participants selected."}
    amount = payload.get("amount", 10)
    send_email = bool(payload.get("send_email", True))
    draft = bool(payload.get("draft", False))

    with IO_LOCK:
        # 1) Bookkeeping first: mark paid in the ledger. If this fails (file
        #    locked), do nothing else - we must never email without recording.
        try:
            updated, missing = write_paid(cids, amount)
        except Exception as e:
            return {"ok": False, "error": str(e)}

        ledger = read_ledger()
        log = read_email_log()

        results = []
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        today = datetime.now().strftime("%Y-%m-%d")
        outlook_broken = None  # cache a hard Outlook failure so we don't retry 300x

        for cid in updated:
            led = ledger.get(cid, {})
            first = led.get("first_name") or ""
            email = (led.get("delivery_email") or "").strip()
            status, err, emailed_at = "skipped", "", ""

            if send_email:
                if not email:
                    status, err = "failed", "no delivery email on file"
                elif outlook_broken is not None:
                    status, err = "failed", outlook_broken
                else:
                    try:
                        subj, body = build_email(first, amount, CONFIG["contact_email"])
                        send_via_outlook(email, subj, body, draft_only=draft)
                        status = "draft" if draft else "sent"
                        emailed_at = now
                    except Exception as e:
                        err = str(e)
                        status = "failed"
                        if "outlook" in err.lower() or "dispatch" in err.lower() or "win32" in err.lower():
                            outlook_broken = err

            prev = log.get(cid, {})
            log[cid] = {
                "cid": cid,
                "first_name": first,
                "delivery_email": email,
                "paid_date": prev.get("paid_date") or today,
                "paid_amount": _fmt_amount(amount),
                "email_status": status if status != "skipped" else prev.get("email_status", "skipped"),
                "emailed_at": emailed_at or prev.get("emailed_at", ""),
                "received_confirmed": prev.get("received_confirmed", ""),
                "received_confirmed_at": prev.get("received_confirmed_at", ""),
                "error": err,
            }
            results.append({"cid": cid, "email_status": status, "error": err})

        write_email_log(log)

    ok_emails = sum(1 for r in results if r["email_status"] in ("sent", "draft"))
    failed = [r for r in results if r["email_status"] == "failed"]
    return {
        "ok": True,
        "marked_paid": len(updated),
        "missing_cids": missing,
        "emails_ok": ok_emails,
        "emails_failed": len(failed),
        "failed": failed[:20],
        "send_email": send_email,
        "draft": draft,
    }


def action_resend(payload):
    """Re-send (or re-draft) the thank-you email to people already marked paid.
    Does NOT touch the ledger — only updates the email log. Used to retry rows
    whose email failed (e.g. Outlook was closed the first time)."""
    cids = [str(c).strip() for c in payload.get("cids", []) if str(c).strip()]
    single = str(payload.get("cid", "")).strip()
    if single and single not in cids:
        cids.append(single)
    if not cids:
        return {"ok": False, "error": "No participants specified."}
    amount = payload.get("amount", 10)
    draft = bool(payload.get("draft", False))
    with IO_LOCK:
        ledger = read_ledger()
        log = read_email_log()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        outlook_broken = None
        results = []
        for cid in cids:
            led = ledger.get(cid, {})
            row = log.get(cid, {})
            first = row.get("first_name") or led.get("first_name") or ""
            email = (row.get("delivery_email") or led.get("delivery_email") or "").strip()
            amt = row.get("paid_amount") or amount
            status, err, emailed_at = "failed", "", ""
            if not email:
                err = "no delivery email on file"
            elif outlook_broken is not None:
                err = outlook_broken
            else:
                try:
                    subj, body = build_email(first, amt, CONFIG["contact_email"])
                    send_via_outlook(email, subj, body, draft_only=draft)
                    status = "draft" if draft else "sent"
                    emailed_at = now
                except Exception as e:
                    err = str(e)
                    if "outlook" in err.lower() or "dispatch" in err.lower() or "win32" in err.lower():
                        outlook_broken = err
            base = dict(row) if row else {
                "cid": cid, "first_name": first, "delivery_email": email,
                "paid_date": "", "paid_amount": _fmt_amount(amt),
                "received_confirmed": "", "received_confirmed_at": "",
            }
            base["cid"] = cid
            base["email_status"] = status
            if emailed_at:
                base["emailed_at"] = emailed_at
            base["error"] = err
            log[cid] = base
            results.append({"cid": cid, "email_status": status, "error": err})
        write_email_log(log)
    ok = sum(1 for r in results if r["email_status"] in ("sent", "draft"))
    fail = [r for r in results if r["email_status"] == "failed"]
    return {"ok": True, "emails_ok": ok, "emails_failed": len(fail), "failed": fail[:20]}


def action_confirm(payload):
    cid = str(payload.get("cid", "")).strip()
    confirmed = bool(payload.get("confirmed", True))
    if not cid:
        return {"ok": False, "error": "Missing cid."}
    with IO_LOCK:
        log = read_email_log()
        if cid not in log:
            return {"ok": False, "error": f"{cid} is not in the payment log."}
        log[cid]["received_confirmed"] = "yes" if confirmed else ""
        log[cid]["received_confirmed_at"] = (
            datetime.now().strftime("%Y-%m-%d %H:%M:%S") if confirmed else "")
        write_email_log(log)
    return {"ok": True, "cid": cid, "confirmed": confirmed}


# -- HTTP server --------------------------------------------------------------
class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass  # quiet

    def _send(self, code, body, ctype="application/json"):
        data = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _json(self, obj, code=200):
        self._send(code, json.dumps(obj), "application/json; charset=utf-8")

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            try:
                html = TEMPLATE_PATH.read_text(encoding="utf-8")
            except Exception as e:
                return self._send(500, f"Template missing: {e}", "text/plain")
            return self._send(200, html, "text/html; charset=utf-8")
        if self.path.startswith("/api/state"):
            try:
                return self._json(build_state())
            except Exception as e:
                return self._json({"error": str(e)}, 500)
        return self._send(404, "not found", "text/plain")

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0) or 0)
        raw = self.rfile.read(length) if length else b"{}"
        try:
            payload = json.loads(raw or b"{}")
        except Exception:
            return self._json({"ok": False, "error": "bad JSON"}, 400)
        try:
            if self.path == "/api/mark_paid":
                return self._json(action_mark_paid(payload))
            if self.path == "/api/resend":
                return self._json(action_resend(payload))
            if self.path == "/api/confirm":
                return self._json(action_confirm(payload))
        except Exception as e:
            return self._json({"ok": False, "error": str(e)}, 500)
        return self._json({"ok": False, "error": "unknown endpoint"}, 404)


def main():
    if openpyxl is None:
        print("openpyxl is required. Install it in the shared venv (see setup.bat).")
    print(f"Payment console - reading {TRACKER_PATH}")
    print(f"Open http://127.0.0.1:{PORT}  (Ctrl+C to stop)")
    try:
        webbrowser.open(f"http://127.0.0.1:{PORT}")
    except Exception:
        pass
    ThreadingHTTPServer(("127.0.0.1", PORT), Handler).serve_forever()


if __name__ == "__main__":
    main()
