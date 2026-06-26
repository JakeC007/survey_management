"""
examine_write.py
------------------
The ONLY part of examine_indv that writes. Everything else is read-only.

mark_fraud(store, response_id, reason) marks a participant as fraud everywhere
the rest of the project looks, so the change is durable and the dashboard picks
it up:

  1. data/fraud_blacklist.csv          <- the SOURCE OF TRUTH. The payment
                                          pipeline rebuilds its fraud verdict
                                          from this file on every run, so an
                                          entry here survives re-runs.
  2. data/payment_tracker.xlsx         <- sets fraud=yes, fraud_reason, and
                                          exclude_recommended=yes on the row
                                          (mirrors what manage_payments writes).
  3. data/payment_report_unpaid.xlsx   <- removes the person from the Pay/Hold
                                          sheets and adds them to the Fraud
                                          sheet, so the dashboard's Fraud count
                                          updates immediately (without a re-run).

Writes are atomic (write a temp file, then replace) and idempotent (re-marking
someone already fraud is a no-op per file). If a workbook is open in Excel the
write fails with a clear message, reported back to the caller.
"""

from __future__ import annotations

import os
import csv
import datetime as dt

import openpyxl

import examine_data as ed

REPORT_XLSX = os.path.join(ed.DATA, "payment_report_unpaid.xlsx")
BLACKLIST_CSV = ed.BLACKLIST_CSV
PAYMENT_XLSX = ed.PAYMENT_XLSX

BLACKLIST_HEADER = ["response_id", "ip", "child_name", "delivery_email", "reasons", "source", "added_at"]


def _now() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def _atomic_save(wb, path: str):
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)


# --------------------------------------------------------------------------- #
# 1. fraud_blacklist.csv
# --------------------------------------------------------------------------- #
def _append_blacklist(rid, ip, child, email, reason) -> bool:
    """Append a manual fraud row. Returns False if the RID is already listed."""
    existing = set()
    if os.path.exists(BLACKLIST_CSV):
        with open(BLACKLIST_CSV, newline="", encoding="utf-8-sig") as f:
            for i, row in enumerate(csv.reader(f)):
                if i == 0 or not row:
                    continue
                existing.add(row[0].strip())
    if rid in existing:
        return False

    new_file = not os.path.exists(BLACKLIST_CSV) or os.path.getsize(BLACKLIST_CSV) == 0
    needs_nl = False
    if not new_file:
        with open(BLACKLIST_CSV, "rb") as f:
            f.seek(-1, os.SEEK_END)
            needs_nl = f.read(1) not in (b"\n", b"\r")

    with open(BLACKLIST_CSV, "a", newline="", encoding="utf-8") as f:
        if needs_nl:
            f.write("\n")
        w = csv.writer(f)
        if new_file:
            w.writerow(BLACKLIST_HEADER)
        w.writerow([rid, ip, child, email, reason, "manual", _now()])
    return True


# --------------------------------------------------------------------------- #
# 2. payment_tracker.xlsx
# --------------------------------------------------------------------------- #
def _update_payment_tracker(rid, reason):
    if not os.path.exists(PAYMENT_XLSX):
        return None, False
    wb = openpyxl.load_workbook(PAYMENT_XLSX)
    ws = wb["Payments"] if "Payments" in wb.sheetnames else wb.worksheets[0]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {name: i for i, name in enumerate(hdr)}
    if "cid" not in col:
        wb.close()
        return None, False
    cid_i = col["cid"]
    target = None
    for row in ws.iter_rows(min_row=2):
        if cid_i < len(row) and str(row[cid_i].value).strip() == rid:
            target = row
            break
    if target is None:
        wb.close()
        return None, False

    # already fraud? no-op (keeps re-marks from stacking reasons)
    if "fraud" in col and str(target[col["fraud"]].value or "").strip().lower() == "yes":
        wb.close()
        return "payment_tracker.xlsx (already fraud=yes)", False

    def setcell(name, value):
        i = col.get(name)
        if i is not None and i < len(target):
            target[i].value = value

    setcell("fraud", "yes")
    existing_reason = str(target[col["fraud_reason"]].value or "").strip() if "fraud_reason" in col else ""
    setcell("fraud_reason", reason if not existing_reason else f"{existing_reason}; {reason}")
    setcell("exclude_recommended", "yes")
    _atomic_save(wb, PAYMENT_XLSX)
    wb.close()
    return "payment_tracker.xlsx (fraud=yes, exclude_recommended=yes)", True


# --------------------------------------------------------------------------- #
# 3. payment_report_unpaid.xlsx
# --------------------------------------------------------------------------- #
def _update_report(rid, first, email, ip, reason, completed_at):
    if not os.path.exists(REPORT_XLSX):
        return [], ["payment_report_unpaid.xlsx (not found - will appear on next pipeline run)"]
    wb = openpyxl.load_workbook(REPORT_XLSX)
    changes, skipped = [], []

    def cid_col(ws):
        hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        return hdr.index("cid") if "cid" in hdr else None

    for sn in wb.sheetnames:
        ws = wb[sn]
        low = sn.lower()
        ci = cid_col(ws)
        if ci is None:
            continue
        if low.startswith("pay") or low.startswith("hold"):
            to_delete = [r[0].row for r in ws.iter_rows(min_row=2)
                         if ci < len(r) and str(r[ci].value).strip() == rid]
            for ridx in sorted(to_delete, reverse=True):
                ws.delete_rows(ridx, 1)
            if to_delete:
                changes.append(f"payment_report_unpaid.xlsx (removed from '{sn}')")
        elif low.startswith("fraud"):
            present = any(ci < len(r) and str(r[ci].value).strip() == rid
                          for r in ws.iter_rows(min_row=2))
            if present:
                skipped.append(f"payment_report_unpaid.xlsx (already in '{sn}')")
            else:
                hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
                values = {
                    "first_name": first, "delivery_email": email, "cid": rid,
                    "survey_ip": ip, "fraud_reason": reason, "survey_completed_at": completed_at,
                }
                ws.append([values.get(h, "") for h in hdr])
                changes.append(f"payment_report_unpaid.xlsx (added to '{sn}')")

    if changes:
        _atomic_save(wb, REPORT_XLSX)
    wb.close()
    return changes, skipped


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #
def mark_fraud(store, response_id, reason: str = "") -> dict:
    rid = ed._s(response_id)
    r = store.tracker_by_rid.get(rid)
    if not r:
        return {"ok": False, "error": f"No participant found for {rid!r}."}

    pay = store.pay_by_cid.get(rid, {})
    reason = ed._s(reason) or "manually marked as fraud via examine_indv"
    child = ed._s(r.get("child_name"))
    email = ed._s(r.get("delivery_email")) or ed._s(pay.get("delivery_email"))
    ip = ed._s(pay.get("survey_ip"))
    first = ed._s(pay.get("first_name")) or (child.split()[0] if child else "")
    completed_at = ed._s(r.get("survey_completed_at")) or ed._s(pay.get("survey_completed_at"))

    changed, skipped, errors = [], [], []

    try:
        if _append_blacklist(rid, ip, child, email, reason):
            changed.append("fraud_blacklist.csv (added)")
        else:
            skipped.append("fraud_blacklist.csv (already listed)")
    except Exception as e:  # noqa: BLE001
        errors.append(f"fraud_blacklist.csv: {e}")

    try:
        msg, did = _update_payment_tracker(rid, reason)
        if msg is None:
            skipped.append("payment_tracker.xlsx (no matching row)")
        else:
            (changed if did else skipped).append(msg)
    except PermissionError:
        errors.append("payment_tracker.xlsx is open in Excel - close it and try again")
    except Exception as e:  # noqa: BLE001
        errors.append(f"payment_tracker.xlsx: {e}")

    try:
        ch, sk = _update_report(rid, first, email, ip, reason, completed_at)
        changed.extend(ch)
        skipped.extend(sk)
    except PermissionError:
        errors.append("payment_report_unpaid.xlsx is open in Excel - close it and try again")
    except Exception as e:  # noqa: BLE001
        errors.append(f"payment_report_unpaid.xlsx: {e}")

    return {"ok": not errors, "rid": rid, "reason": reason,
            "changed": changed, "skipped": skipped, "errors": errors}
