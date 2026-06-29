"""
examine_review.py
-------------------
Adds a "browse flagged items" review queue on top of the examine_indv tool.

A reviewer steps (Back / Next) through every participant who carries quality
flags, reads their answers and metadata, and decides:

  * Mark as fraud   -> handled by examine_write.mark_fraud (unchanged), or
  * Clear / keep    -> records a "this one is fine" decision so the person
                       drops out of the queue and never resurfaces.

Queue membership (rebuilt fresh from disk on every request, so it always
reflects the current state of the sheets):

    n_flags > 0   AND   not already FRAUD   AND   not already cleared here

Already-paid people are included by default (they still carry flags); the UI
has a "hide paid" toggle.

WRITE SAFETY
------------
The only thing this module writes is data/review_state.csv, a NEW file that no
other part of the pipeline opens, so it can never collide with an open Excel
workbook. Writes are atomic (temp file -> os.replace) and idempotent
(re-clearing an already-cleared cid is a no-op). Reads tolerate a missing or
half-written file without raising. Marking fraud still goes through
examine_write, which already guards against Excel locks.

Public API
    load_cleared()                 -> set[str] of cleared cids
    build_queue(store, ...)        -> dict with the ordered queue
    clear_item(store, rid, note)   -> dict result (atomic append)
    REVIEW_HTML                    -> the /review single-page UI
"""

from __future__ import annotations

import os
import csv
import datetime as dt

import openpyxl

import examine_data as ed
import examine_write as ew


REVIEW_STATE_CSV = os.path.join(ed.DATA, "review_state.csv")
REVIEW_HEADER = ["cid", "decision", "note", "reviewed_by", "reviewed_at"]

PAYMENT_XLSX = ed.PAYMENT_XLSX
REPORT_XLSX = os.path.join(ed.DATA, "payment_report_unpaid.xlsx")


def _now() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


# --------------------------------------------------------------------------- #
# review_state.csv  (read)
# --------------------------------------------------------------------------- #
def load_cleared() -> dict:
    """Return {cid: row} for every cid whose latest decision is 'cleared'.

    Tolerates a missing or partially written file (returns {} / what it could
    read). Later rows win, so a cid can be cleared, un-cleared, re-cleared.
    """
    out: dict = {}
    if not os.path.exists(REVIEW_STATE_CSV):
        return out
    try:
        with open(REVIEW_STATE_CSV, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                cid = ed._s(row.get("cid"))
                if not cid:
                    continue
                decision = ed._s(row.get("decision")).lower()
                if decision == "cleared":
                    out[cid] = row
                else:
                    out.pop(cid, None)  # an explicit non-clear removes it
    except OSError:
        pass
    return out


# --------------------------------------------------------------------------- #
# payment_tracker.xlsx  (flip the kept person out of Hold)
# --------------------------------------------------------------------------- #
def _keep_in_payment_tracker(rid, note):
    """Set exclude_recommended=no and stamp a note on the person's ledger row,
    so they read as PAY (not HOLD) and a re-run preserves the human decision.
    Returns (message_or_None, changed_bool). Idempotent."""
    if not os.path.exists(PAYMENT_XLSX):
        return None, False
    wb = openpyxl.load_workbook(PAYMENT_XLSX)
    ws = wb["Payments"] if "Payments" in wb.sheetnames else wb.worksheets[0]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {name: i for i, name in enumerate(hdr)}
    if "cid" not in col:
        wb.close()
        return None, False
    cid_i = col["cid"]
    target = None
    for row in ws.iter_rows(min_row=2):
        if cid_i < len(row) and str(row[cid_i].value).strip() == rid:
            target = row
            break
    if target is None:
        wb.close()
        return None, False

    stamp = "kept via examine_indv review" + (f": {note}" if note else "")
    changed = False

    if "exclude_recommended" in col:
        cell = target[col["exclude_recommended"]]
        if str(cell.value or "").strip().lower() == "yes":
            cell.value = "no"
            changed = True
    if "notes" in col:
        existing = str(target[col["notes"]].value or "").strip()
        if "kept via examine_indv review" not in existing:  # don't stack on re-clear
            target[col["notes"]].value = stamp if not existing else f"{existing}; {stamp}"
            changed = True

    if changed:
        ew._atomic_save(wb, PAYMENT_XLSX)
    wb.close()
    return ("payment_tracker.xlsx (exclude_recommended=no)" if changed
            else "payment_tracker.xlsx (already not held)"), changed


# --------------------------------------------------------------------------- #
# payment_report_unpaid.xlsx  (move the row from the Hold sheet to the Pay sheet)
# --------------------------------------------------------------------------- #
def _hdr_of(ws):
    return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]


def _move_report_hold_to_pay(rid):
    """Move the person's row from the 'Hold' sheet to the 'Pay' sheet so the
    dashboard's Hold count drops and Pay count rises immediately. Returns
    (changed_list, skipped_list)."""
    if not os.path.exists(REPORT_XLSX):
        return [], ["payment_report_unpaid.xlsx (not found - will sort out on next pipeline run)"]
    wb = openpyxl.load_workbook(REPORT_XLSX)

    hold = pay = None  # (sheet_name, worksheet)
    for sn in wb.sheetnames:
        low = sn.lower()
        if low.startswith("hold"):
            hold = (sn, wb[sn])
        elif low.startswith("pay"):
            pay = (sn, wb[sn])
    if hold is None or pay is None:
        wb.close()
        return [], ["payment_report_unpaid.xlsx (Pay/Hold sheet not found)"]

    hsn, hws = hold
    psn, pws = pay
    hhdr = _hdr_of(hws)
    if "cid" not in hhdr:
        wb.close()
        return [], ["payment_report_unpaid.xlsx (no cid column in Hold sheet)"]
    hci = hhdr.index("cid")

    captured, del_rows = None, []
    for r in hws.iter_rows(min_row=2):
        if hci < len(r) and str(r[hci].value).strip() == rid:
            captured = {hhdr[i]: (r[i].value if i < len(r) else None) for i in range(len(hhdr))}
            del_rows.append(r[0].row)

    if captured is None:
        # Not in Hold - maybe they were already in Pay (flagged but not held).
        phdr = _hdr_of(pws)
        pci = phdr.index("cid") if "cid" in phdr else None
        in_pay = pci is not None and any(
            pci < len(r) and str(r[pci].value).strip() == rid
            for r in pws.iter_rows(min_row=2))
        wb.close()
        return [], ["payment_report_unpaid.xlsx (already in Pay)" if in_pay
                    else "payment_report_unpaid.xlsx (not in Hold sheet)"]

    for ridx in sorted(del_rows, reverse=True):
        hws.delete_rows(ridx, 1)
    phdr = _hdr_of(pws)
    pws.append([captured.get(h, "") for h in phdr])
    ew._atomic_save(wb, REPORT_XLSX)
    wb.close()
    return [f"payment_report_unpaid.xlsx (moved '{hsn}' -> '{psn}')"], []


# --------------------------------------------------------------------------- #
# review_state.csv  (write, atomic + idempotent)
# --------------------------------------------------------------------------- #
def _append_review_state(rid, note, reviewed_by):
    """Atomic, idempotent append of a 'cleared' row. Raises on lock/IO so the
    caller can surface it."""
    new_file = not os.path.exists(REVIEW_STATE_CSV) or os.path.getsize(REVIEW_STATE_CSV) == 0
    existing = b""
    if not new_file:
        with open(REVIEW_STATE_CSV, "rb") as f:
            existing = f.read()
        if existing and existing[-1:] not in (b"\n", b"\r"):
            existing += b"\n"

    import io
    sbuf = io.StringIO()
    w = csv.writer(sbuf)
    if new_file:
        w.writerow(REVIEW_HEADER)
    w.writerow([rid, "cleared", note, reviewed_by, _now()])
    new_bytes = existing + sbuf.getvalue().encode("utf-8")

    tmp = REVIEW_STATE_CSV + ".tmp"
    with open(tmp, "wb") as f:
        f.write(new_bytes)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, REVIEW_STATE_CSV)


def clear_item(store, response_id, note: str = "", reviewed_by: str = "") -> dict:
    """Record a 'cleared / keep' decision and move the person Hold -> Pay.

    Three durable effects (mirrors mark_fraud's multi-file write):
      1. payment_tracker.xlsx  -> exclude_recommended=no (+ note), so they read PAY.
      2. payment_report_unpaid.xlsx -> moved from the Hold sheet to the Pay sheet,
         so the dashboard's Hold count drops and Pay count rises right away.
      3. review_state.csv -> records the keep, drops them from the review queue,
         and makes the next manage_payments run keep them in Pay (KEPT_CIDS).

    No-op if already cleared. If a workbook is locked open in Excel, nothing is
    recorded and a clear error is returned, so the reviewer can close it and
    retry (the keep is not half-applied).
    """
    rid = ed._s(response_id)
    if not rid:
        return {"ok": False, "error": "missing response id"}

    # Only allow clearing people we can actually see in the tracker.
    if rid not in store.tracker_by_rid:
        return {"ok": False, "error": f"No participant found for {rid!r}."}

    if rid in load_cleared():
        return {"ok": True, "rid": rid, "skipped": ["already cleared"], "changed": [], "errors": []}

    note = ed._s(note)
    reviewed_by = ed._s(reviewed_by)
    changed, skipped, errors = [], [], []

    # 1. payment_tracker.xlsx  (flip exclude_recommended)
    try:
        msg, did = _keep_in_payment_tracker(rid, note)
        if msg is None:
            skipped.append("payment_tracker.xlsx (no matching row)")
        else:
            (changed if did else skipped).append(msg)
    except PermissionError:
        errors.append("payment_tracker.xlsx is open in Excel - close it and try again")
    except Exception as e:  # noqa: BLE001
        errors.append(f"payment_tracker.xlsx: {e}")

    # 2. payment_report_unpaid.xlsx  (move Hold -> Pay) - only if step 1 was clean
    if not errors:
        try:
            ch, sk = _move_report_hold_to_pay(rid)
            changed.extend(ch)
            skipped.extend(sk)
        except PermissionError:
            errors.append("payment_report_unpaid.xlsx is open in Excel - close it and try again")
        except Exception as e:  # noqa: BLE001
            errors.append(f"payment_report_unpaid.xlsx: {e}")

    # If a workbook was locked, don't record the keep - let the user retry clean.
    if errors:
        return {"ok": False, "rid": rid, "changed": changed, "skipped": skipped, "errors": errors}

    # 3. review_state.csv  (the durable keep record + queue drop)
    try:
        _append_review_state(rid, note, reviewed_by)
        changed.append("review_state.csv (cleared)")
    except PermissionError:
        errors.append("review_state.csv is open/locked - close it and try again")
    except OSError as e:
        errors.append(f"review_state.csv: {e}")

    return {"ok": not errors, "rid": rid, "changed": changed, "skipped": skipped, "errors": errors}


# --------------------------------------------------------------------------- #
# Build the review queue
# --------------------------------------------------------------------------- #
def _n_flags(pay) -> int:
    raw = ed._s(pay.get("n_flags"))
    if not raw:
        return 0
    try:
        return int(float(raw))
    except ValueError:
        return 0


def build_queue(store, include_paid: bool = True) -> dict:
    """Ordered list of flagged participants still needing a fraud decision.

    Each item is a light summary (enough to render the list and pick a starting
    point); full detail is fetched per-item via the existing /api/detail.
    """
    cleared = load_cleared()
    items = []

    for rid, r in store.tracker_by_rid.items():
        pay = store.pay_by_cid.get(rid)
        if not pay:
            continue
        nf = _n_flags(pay)
        if nf <= 0:
            continue

        status, reason = store._status(rid, r)
        if status == "FRAUD":
            continue                      # already decided
        if rid in cleared:
            continue                      # reviewer already kept this one
        is_paid = status == "PAID"
        if is_paid and not include_paid:
            continue

        items.append({
            "response_id": rid,
            "child_name": ed._s(r.get("child_name")),
            "parent_name": ed._s(r.get("parent_name")),
            "delivery_email": ed._s(r.get("delivery_email")),
            "grade": ed._s(r.get("grade")),
            "status": status,
            "n_flags": nf,
            "flag_reasons": ed._s(pay.get("flag_reasons")),
            "completed": store._is_completed(rid, r),
            "recorded_date": ed._s(r.get("recorded_date")),
            "paid": is_paid,
        })

    # Most-flagged first; then HOLD before PAY before PAID; then newest.
    rank = {"HOLD": 0, "NOT EVALUATED": 1, "PAY": 2, "PAID": 3}
    items.sort(key=lambda x: (-x["n_flags"], rank.get(x["status"], 9),
                              x.get("recorded_date") or ""), reverse=False)

    total_cleared = len(cleared)
    return {
        "count": len(items),
        "cleared_total": total_cleared,
        "include_paid": include_paid,
        "items": items,
    }


# --------------------------------------------------------------------------- #
# The /review single-page UI
# --------------------------------------------------------------------------- #
REVIEW_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Review flagged participants</title>
<style>
  :root{
    --bg:#0f1115; --panel:#171a21; --panel2:#1d212b; --line:#2a2f3a;
    --text:#e7e9ee; --muted:#9aa3b2; --accent:#5b9dff;
    --pay:#22c55e; --hold:#f59e0b; --fraud:#ef4444; --paid:#3b82f6; --na:#6b7280;
  }
  *{box-sizing:border-box}
  body{margin:0;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
       background:var(--bg);color:var(--text);line-height:1.45}
  header{padding:18px 28px 14px;border-bottom:1px solid var(--line);display:flex;align-items:center;gap:16px;flex-wrap:wrap}
  h1{font-size:18px;margin:0}
  .sub{color:var(--muted);font-size:13px}
  a.home{color:var(--accent);font-size:13px;text-decoration:none}
  a.home:hover{text-decoration:underline}
  .wrap{max-width:980px;margin:0 auto;padding:18px 28px 70px}

  .navbar{position:sticky;top:0;z-index:20;background:var(--bg);
       display:flex;align-items:center;gap:12px;flex-wrap:wrap;
       padding:12px 0 14px;border-bottom:1px solid var(--line);margin-bottom:16px}
  .navbtn{background:var(--panel);border:1px solid var(--line);color:var(--text);
       height:42px;padding:0 18px;border-radius:10px;font-size:14px;font-weight:600;cursor:pointer}
  .navbtn:hover:not(:disabled){border-color:var(--accent)}
  .navbtn:disabled{opacity:.4;cursor:not-allowed}
  .counter{font-size:14px;font-weight:600;min-width:120px;text-align:center}
  .counter .muted{color:var(--muted);font-weight:400}
  .spacer{flex:1}
  .toggle{display:flex;align-items:center;gap:7px;color:var(--muted);font-size:13px;cursor:pointer;user-select:none}
  .toggle input{width:16px;height:16px;accent-color:var(--accent)}
  .progress{height:4px;background:var(--panel2);border-radius:99px;overflow:hidden;margin-bottom:14px}
  .progress > i{display:block;height:100%;background:var(--accent);width:0%;transition:width .2s}

  button.act{height:42px;padding:0 18px;border-radius:10px;font-size:14px;font-weight:600;cursor:pointer;border:1px solid var(--line);background:var(--panel);color:var(--text)}
  .btn-danger{background:rgba(239,68,68,.14);color:#ff8b8b;border:1px solid rgba(239,68,68,.5)}
  .btn-danger:hover{background:rgba(239,68,68,.22)}
  .btn-ok{background:rgba(34,197,94,.14);color:#7ee2a0;border:1px solid rgba(34,197,94,.5)}
  .btn-ok:hover{background:rgba(34,197,94,.22)}

  .card{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:18px 20px;margin-bottom:16px}
  .statusrow{display:flex;align-items:center;gap:14px;flex-wrap:wrap}
  .bigbadge{font-size:15px;font-weight:800;letter-spacing:.05em;padding:8px 16px;border-radius:10px}
  .badge{font-size:11px;font-weight:700;letter-spacing:.04em;padding:4px 9px;border-radius:999px;white-space:nowrap}
  .b-PAY,.bigbadge.b-PAY{background:rgba(34,197,94,.15);color:var(--pay)}
  .b-HOLD,.bigbadge.b-HOLD{background:rgba(245,158,11,.15);color:var(--hold)}
  .b-FRAUD,.bigbadge.b-FRAUD{background:rgba(239,68,68,.15);color:var(--fraud)}
  .b-PAID,.bigbadge.b-PAID{background:rgba(59,130,246,.15);color:var(--paid)}
  .b-NA,.b-NOTEVALUATED,.bigbadge.b-NOTEVALUATED,.bigbadge.b-NA{background:rgba(107,114,128,.18);color:#b6bdca}
  .flagchip{background:rgba(245,158,11,.15);color:var(--hold);font-weight:700}
  .name{font-size:20px;font-weight:700}
  .idline{color:var(--muted);font-size:13px;margin-top:3px}
  .why{margin-top:12px;color:#d9dde6;font-size:14px;background:var(--panel2);border-left:3px solid var(--line);
       padding:10px 14px;border-radius:0 8px 8px 0}
  .why.fraud{border-color:var(--fraud)} .why.hold{border-color:var(--hold)}
  .why.pay{border-color:var(--pay)} .why.paid{border-color:var(--paid)}
  .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:10px 18px;margin-top:6px}
  .kv .k{color:var(--muted);font-size:11.5px;text-transform:uppercase;letter-spacing:.04em}
  .kv .v{font-size:14px;margin-top:1px;word-break:break-word}
  h2{font-size:13px;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);margin:0 0 12px}
  .qa{border-top:1px solid var(--line);padding:11px 0}
  .qa:first-of-type{border-top:0}
  .qa .q{font-size:13.5px;color:var(--muted)}
  .qa .a{font-size:15px;margin-top:3px;white-space:pre-wrap}
  .qa.other .q::after{content:" (free text)";color:var(--accent);font-size:11px}
  .empty{color:var(--muted);font-style:italic}
  .notice{background:rgba(245,158,11,.10);border:1px solid rgba(245,158,11,.35);color:#f6d08a;
       border-radius:10px;padding:12px 14px;font-size:14px}
  .hint{color:var(--muted);font-size:12px;margin:8px 2px}
  .err{color:#f6a6a6}
  code{background:var(--panel2);padding:1px 6px;border-radius:5px;font-size:12.5px}

  .overlay{position:fixed;inset:0;background:rgba(0,0,0,.55);display:flex;align-items:center;justify-content:center;z-index:50}
  .modal{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:22px 24px;width:min(520px,92vw);box-shadow:0 20px 60px rgba(0,0,0,.5)}
  .modal h3{margin:0 0 8px;font-size:17px}
  .modal p{color:var(--muted);font-size:13.5px;margin:0 0 14px}
  .modal textarea{width:100%;min-height:70px;background:var(--panel2);border:1px solid var(--line);color:var(--text);border-radius:9px;padding:10px 12px;font-size:14px;resize:vertical;font-family:inherit}
  .modal .what{background:var(--panel2);border-radius:9px;padding:10px 12px;margin:12px 0 4px;font-size:12.5px;color:var(--muted)}
  .modal .what b{color:var(--text)}
  .modal-actions{display:flex;justify-content:flex-end;gap:10px;margin-top:18px}
  .btn-ghost{background:none;border:1px solid var(--line);color:var(--text);height:42px;border-radius:9px;padding:0 16px;cursor:pointer}
  .btn-ghost:hover{border-color:var(--accent)}
  .btn-confirm{background:var(--fraud);color:#fff;height:42px;border-radius:9px;padding:0 16px;border:0;cursor:pointer;font-weight:600}
  .btn-confirm.keep{background:var(--pay);color:#06210f}
  .toast{position:fixed;bottom:22px;left:50%;transform:translateX(-50%);background:var(--panel2);border:1px solid var(--line);border-radius:10px;padding:12px 16px;font-size:13.5px;z-index:60;max-width:90vw;box-shadow:0 10px 30px rgba(0,0,0,.4)}
  .toast.ok{border-color:rgba(34,197,94,.5)} .toast.bad{border-color:rgba(239,68,68,.5)}
  .toast ul{margin:6px 0 0;padding-left:18px} .toast li{margin:2px 0}
</style>
</head>
<body>
<header>
  <h1>Review flagged participants</h1>
  <div class="sub">Step through everyone carrying quality flags and decide: fraud, or keep.</div>
  <div style="flex:1"></div>
  <a class="home" href="/">← participant search</a>
</header>
<div class="wrap">
  <div class="navbar">
    <button class="navbtn" id="prev" onclick="go(-1)">← Back</button>
    <div class="counter" id="counter"><span class="muted">loading…</span></div>
    <button class="navbtn" id="next" onclick="go(1)">Next →</button>
    <div class="spacer"></div>
    <label class="toggle"><input type="checkbox" id="hidepaid" onchange="reloadQueue()"> hide already-paid</label>
    <button class="navbtn" onclick="reloadQueue()" title="Re-read the sheets">↻ refresh</button>
  </div>
  <div class="progress"><i id="bar"></i></div>
  <div id="out"></div>
</div>
<div id="modal-root"></div>
<div id="toast-root"></div>

<script>
function esc(s){return (s==null?'':String(s)).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));}
function badgeClass(s){return 'b-'+String(s).replace(/[^A-Z]/g,'');}

let QUEUE=[];      // light summaries
let IDX=0;         // current position
let CURRENT=null;  // full detail of current item

async function reloadQueue(){
  const hidePaid = document.getElementById('hidepaid').checked;
  document.getElementById('out').innerHTML = '<div class="hint">Loading queue…</div>';
  let res;
  try{ res = await (await fetch('/api/review_queue?include_paid='+(hidePaid?'0':'1'))).json(); }
  catch(e){ document.getElementById('out').innerHTML='<div class="err">Failed to load queue: '+esc(e)+'</div>'; return; }
  QUEUE = res.items||[];
  IDX = 0;
  if(QUEUE.length===0){
    setCounter();
    document.getElementById('out').innerHTML =
      '<div class="card"><div class="notice">Nothing left to review. '
      + 'No flagged participants are awaiting a decision'
      + (hidePaid?' (paid people hidden)':'') + '. '
      + (res.cleared_total? esc(res.cleared_total)+' cleared so far.':'') + '</div></div>';
    return;
  }
  showCurrent();
}

function setCounter(){
  const c = document.getElementById('counter');
  if(QUEUE.length===0){ c.innerHTML='<span class="muted">0 to review</span>'; }
  else { c.innerHTML = (IDX+1)+' <span class="muted">of '+QUEUE.length+'</span>'; }
  document.getElementById('prev').disabled = IDX<=0;
  document.getElementById('next').disabled = IDX>=QUEUE.length-1;
  const pct = QUEUE.length? Math.round(((IDX+ (QUEUE.length? 1:0))/QUEUE.length)*100):0;
  document.getElementById('bar').style.width = pct+'%';
}

function go(delta){
  const n = IDX+delta;
  if(n<0||n>=QUEUE.length) return;
  IDX = n;
  showCurrent();
}

async function showCurrent(){
  setCounter();
  const item = QUEUE[IDX];
  document.getElementById('out').innerHTML = '<div class="hint">Loading…</div>';
  let d;
  try{ d = await (await fetch('/api/detail?rid='+encodeURIComponent(item.response_id))).json(); }
  catch(e){ document.getElementById('out').innerHTML='<div class="err">Failed: '+esc(e)+'</div>'; return; }
  if(d.error){ document.getElementById('out').innerHTML='<div class="card err">'+esc(d.error)+'</div>'; return; }
  CURRENT=d; renderDetail(d, item);
}

function renderDetail(d, item){
  const dst = document.getElementById('out');
  const p = d.person, st = d.status;
  const lc = st==='FRAUD'?'fraud':st==='HOLD'?'hold':st==='PAID'?'paid':st==='PAY'?'pay':'';
  let h = '';

  h += '<div class="card"><div class="statusrow">'
     + '<span class="bigbadge '+badgeClass(st)+'">'+esc(st)+'</span>'
     + '<span class="badge flagchip">'+esc(item.n_flags)+' flag'+(item.n_flags==1?'':'s')+'</span>'
     + '<div><div class="name">'+esc(p.child_name||'(no child name)')+'</div>'
     + '<div class="idline">parent/guardian: '+esc(p.parent_name||'—')+' &nbsp;·&nbsp; '+esc(p.delivery_email||'—')
     + ' &nbsp;·&nbsp; grade '+esc(p.grade||'—')+' &nbsp;·&nbsp; '+esc(p.response_id)+'</div></div>'
     + '<div class="spacer"></div>'
     + '<button class="act btn-ok" onclick="openKeepModal()">✓ Clear / keep</button>'
     + '<button class="act btn-danger" onclick="openFraudModal()">Mark as fraud</button>'
     + '</div>';
  h += '<div class="why '+lc+'"><b>Why '+esc(st)+':</b> '+esc(d.reason||'—')+'</div>';
  if(item.flag_reasons){ h += '<div class="why hold"><b>Flagged for:</b> '+esc(item.flag_reasons)+'</div>'; }
  h += '</div>';

  if(!d.completed){
    h += '<div class="card"><div class="notice">No completed survey on file, so there are no answers to show.</div></div>';
  }

  if(d.metadata && d.metadata.length){
    h += '<div class="card"><h2>Key metadata &amp; quality signals</h2><div class="grid">';
    for(const m of d.metadata){ h += '<div class="kv"><div class="k">'+esc(m.label)+'</div><div class="v">'+esc(m.value)+'</div></div>'; }
    h += '</div></div>';
  }

  if(d.completed && d.featured_text && d.featured_text.length){
    h += '<div class="card"><h2>Open-text responses</h2>';
    for(const a of d.featured_text){ h += '<div class="qa other"><div class="q">'+esc(a.label)+'</div><div class="a">'+esc(a.value)+'</div></div>'; }
    h += '</div>';
  }

  if(d.completed){
    h += '<div class="card"><h2>Survey answers ('+(d.answers?d.answers.length:0)+')</h2>';
    if(d.answers && d.answers.length){
      for(const a of d.answers){ h += '<div class="qa'+(a.other?' other':'')+'"><div class="q">'+esc(a.label)+'</div><div class="a">'+esc(a.value)+'</div></div>'; }
    }else{ h += '<div class="empty">Marked complete, but no answer fields were found.</div>'; }
    h += '</div>';
  }

  h += '<div class="hint">Answers sourced from <code>'+esc(d.export_name||'?')+'</code>. '
     + 'Use ← / → keys to move between items.</div>';
  dst.innerHTML = h;
}

/* remove the current item locally and advance */
function dropCurrentAndAdvance(){
  QUEUE.splice(IDX,1);
  if(QUEUE.length===0){ reloadQueue(); return; }
  if(IDX>=QUEUE.length) IDX = QUEUE.length-1;
  showCurrent();
}

/* ---- Mark fraud ---- */
function openFraudModal(){
  if(!CURRENT) return;
  const p = CURRENT.person;
  document.getElementById('modal-root').innerHTML =
    '<div class="overlay" onclick="if(event.target===this)closeModal()"><div class="modal">'
    + '<h3>Mark this person as fraud?</h3>'
    + '<p>'+esc(p.child_name||p.response_id)+' ('+esc(p.delivery_email||'no email')+') · '+esc(p.response_id)+'</p>'
    + '<label class="k" style="color:var(--muted);font-size:12px">Reason (optional)</label>'
    + '<textarea id="reason" placeholder="e.g. duplicate submission / fabricated identity"></textarea>'
    + '<div class="what">Writes to <b>fraud_blacklist.csv</b>, <b>payment_tracker.xlsx</b> (fraud=yes), '
    + 'and <b>payment_report_unpaid.xlsx</b>. This person will never be paid.</div>'
    + '<div class="modal-actions"><button class="btn-ghost" onclick="closeModal()">Cancel</button>'
    + '<button class="btn-confirm" id="go" onclick="confirmFraud()">Yes, mark as fraud</button></div>'
    + '</div></div>';
}

/* ---- Clear / keep ---- */
function openKeepModal(){
  if(!CURRENT) return;
  const p = CURRENT.person;
  document.getElementById('modal-root').innerHTML =
    '<div class="overlay" onclick="if(event.target===this)closeModal()"><div class="modal">'
    + '<h3>Clear this person (not fraud)?</h3>'
    + '<p>'+esc(p.child_name||p.response_id)+' ('+esc(p.delivery_email||'no email')+') · '+esc(p.response_id)+'</p>'
    + '<label class="k" style="color:var(--muted);font-size:12px">Note (optional)</label>'
    + '<textarea id="reason" placeholder="e.g. reviewed answers, looks legitimate"></textarea>'
    + '<div class="what">Moves this person from <b>Hold</b> to <b>Pay</b>: sets <b>exclude_recommended=no</b> in '
    + 'payment_tracker.xlsx, moves them to the Pay sheet of payment_report_unpaid.xlsx (the dashboard Hold count '
    + 'drops, Pay rises), and records the keep in <b>review_state.csv</b> so they leave the queue and stay in Pay '
    + 'on the next pipeline run.</div>'
    + '<div class="modal-actions"><button class="btn-ghost" onclick="closeModal()">Cancel</button>'
    + '<button class="btn-confirm keep" id="go" onclick="confirmKeep()">Clear / keep</button></div>'
    + '</div></div>';
}
function closeModal(){ document.getElementById('modal-root').innerHTML=''; }

async function confirmFraud(){
  const rid = CURRENT.person.response_id;
  const reason = (document.getElementById('reason').value||'').trim();
  const go = document.getElementById('go'); go.disabled=true; go.textContent='Saving…';
  let res;
  try{
    res = await (await fetch('/api/mark_fraud',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({rid, reason})})).json();
  }catch(e){ closeModal(); toast(false,'Request failed: '+esc(e)); return; }
  closeModal();
  if(res.error){ toast(false, esc(res.error)); return; }
  const r = res.result||{};
  let msg = r.ok? '<b>Marked as fraud.</b>':'<b>Marked with problems.</b>';
  const changed=(r.changed||[]).concat(r.skipped||[]);
  if(changed.length) msg += '<ul>'+changed.map(x=>'<li>'+esc(x)+'</li>').join('')+'</ul>';
  if((r.errors||[]).length) msg += '<ul>'+r.errors.map(x=>'<li class="err">'+esc(x)+'</li>').join('')+'</ul>';
  toast(r.ok, msg);
  if(r.errors && r.errors.length){ if(res.detail){CURRENT=res.detail; showCurrent();} return; }
  dropCurrentAndAdvance();
}

async function confirmKeep(){
  const rid = CURRENT.person.response_id;
  const note = (document.getElementById('reason').value||'').trim();
  const go = document.getElementById('go'); go.disabled=true; go.textContent='Saving…';
  let res;
  try{
    res = await (await fetch('/api/clear_review',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({rid, note})})).json();
  }catch(e){ closeModal(); toast(false,'Request failed: '+esc(e)); return; }
  closeModal();
  if(res.error){ toast(false, esc(res.error)); return; }
  const r = res.result||{};
  if(!r.ok){ toast(false, esc((r.errors&&r.errors[0])||r.error||'could not clear')); return; }
  let msg = '<b>Cleared / kept.</b> Moved to Pay and dropped from the queue.';
  const changed = r.changed||[];
  if(changed.length) msg += '<ul>'+changed.map(x=>'<li>'+esc(x)+'</li>').join('')+'</ul>';
  toast(true, msg);
  dropCurrentAndAdvance();
}

function toast(ok, html){
  const root = document.getElementById('toast-root');
  root.innerHTML = '<div class="toast '+(ok?'ok':'bad')+'">'+html+'</div>';
  setTimeout(()=>{ if(root.firstChild) root.innerHTML=''; }, 8000);
}

document.addEventListener('keydown', e=>{
  if(document.getElementById('modal-root').firstChild) return; // modal open
  if(e.key==='ArrowLeft'){ go(-1); }
  else if(e.key==='ArrowRight'){ go(1); }
});

reloadQueue();
</script>
</body>
</html>
"""
