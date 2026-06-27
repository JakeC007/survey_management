"""
examine_data.py
-----------------
Read-only data layer for the "examine an individual participant" tool.

Loads, from the survey_management project:
  - data/participant_tracker_auto.xlsx   (the master roster / lookup table)
  - data/payment_tracker.xlsx            (pay / hold / fraud / paid decision + reasons)
  - data/fraud_blacklist.csv             (cross-check for fraud by RID / email / IP)
  - ingest/K12 Privacy and AI Extension_*.zip  (the actual survey answers; latest export)

Public API:
  load_all()                       -> Store  (call once at startup; re-call to refresh)
  Store.search(query)              -> dict with candidate list (handles name collisions)
  Store.detail(response_id)        -> full record: status, reason, answers, metadata

The join key across every file is the consent ResponseId, which equals:
    tracker.response_id  ==  payment_tracker.cid  ==  survey_export "cid" column

The data/ and ingest/ directories can be overridden with the environment
variables EXAMINE_DATA_DIR and EXAMINE_INGEST_DIR (used for testing).
"""

from __future__ import annotations

import csv
import io
import os
import re
import glob
import zipfile
import datetime as dt

import openpyxl


# --------------------------------------------------------------------------- #
# Paths  (data/ and ingest/ live one level up, in the repo root)
# --------------------------------------------------------------------------- #
HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
DATA = os.environ.get("EXAMINE_DATA_DIR") or os.path.join(REPO, "data")
INGEST = os.environ.get("EXAMINE_INGEST_DIR") or os.path.join(REPO, "ingest")

TRACKER_XLSX = os.path.join(DATA, "participant_tracker_auto.xlsx")
PAYMENT_XLSX = os.path.join(DATA, "payment_tracker.xlsx")
BLACKLIST_CSV = os.path.join(DATA, "fraud_blacklist.csv")

RID_RE = re.compile(r"^R_[A-Za-z0-9]+$")
# Long free-text "walk us through your decision" questions, surfaced in their
# own card above the regular survey answers (and kept out of that list).
FEATURED_TEXT_CODES = ("Q62", "Q63")
# Qualtrics timing sub-columns we never want to show as a "question"
TIMING_BITS = ("First Click", "Last Click", "Page Submit", "Click Count")


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def _s(v) -> str:
    """Normalise a cell to a clean string ('' for None)."""
    if v is None:
        return ""
    return str(v).strip()


def _norm(v) -> str:
    return " ".join(_s(v).lower().split())


def _yes(v) -> bool:
    return _s(v).lower() in ("yes", "true", "1", "y")


def _read_xlsx(path: str, sheet: str | None = None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = [_s(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append({header[i]: r[i] for i in range(len(header))})
    return header, out


# --------------------------------------------------------------------------- #
# Offline ZIP-code -> "City, ST" annotation
# ---------------------------------------------------------------------------
# Keyed off a small bundled file (zip_lookup.csv, indexed by the 3-digit ZIP
# prefix) so it works with NO internet connection. If that file is missing or a
# prefix isn't in it, we fall back to a built-in prefix->state table, and
# finally to nothing (the annotation is simply skipped). Nothing here raises,
# so a missing/garbled file can never break the viewer.
# --------------------------------------------------------------------------- #
ZIP_LOOKUP_CSV = os.path.join(HERE, "zip_lookup.csv")

# ZIP3 prefix ranges -> US state. Only used when a prefix is absent from
# zip_lookup.csv, to still surface at least the state for an unseen ZIP.
_STATE_PREFIX_RANGES = [
    (6, 9, "PR"), (10, 27, "MA"), (28, 29, "RI"), (30, 38, "NH"),
    (39, 49, "ME"), (50, 54, "VT"), (55, 55, "MA"), (56, 59, "VT"),
    (60, 69, "CT"), (70, 89, "NJ"), (100, 149, "NY"), (150, 196, "PA"),
    (197, 199, "DE"), (200, 205, "DC"), (206, 219, "MD"), (220, 246, "VA"),
    (247, 268, "WV"), (270, 289, "NC"), (290, 299, "SC"), (300, 319, "GA"),
    (320, 349, "FL"), (350, 369, "AL"), (370, 385, "TN"), (386, 397, "MS"),
    (398, 399, "GA"), (400, 427, "KY"), (430, 459, "OH"), (460, 479, "IN"),
    (480, 499, "MI"), (500, 528, "IA"), (530, 549, "WI"), (550, 567, "MN"),
    (570, 577, "SD"), (580, 588, "ND"), (590, 599, "MT"), (600, 629, "IL"),
    (630, 658, "MO"), (660, 679, "KS"), (680, 693, "NE"), (700, 714, "LA"),
    (716, 729, "AR"), (730, 749, "OK"), (750, 799, "TX"), (800, 816, "CO"),
    (820, 831, "WY"), (832, 838, "ID"), (840, 847, "UT"), (850, 865, "AZ"),
    (870, 884, "NM"), (885, 885, "TX"), (889, 898, "NV"), (900, 961, "CA"),
    (967, 968, "HI"), (970, 979, "OR"), (980, 994, "WA"), (995, 999, "AK"),
]

_ZIP3_CITY_STATE = None  # lazy cache: {prefix: (city, state)}


def _load_zip_lookup() -> dict:
    """Load zip_lookup.csv into {prefix: (city, state)}; '' table if absent."""
    global _ZIP3_CITY_STATE
    if _ZIP3_CITY_STATE is not None:
        return _ZIP3_CITY_STATE
    table: dict = {}
    try:
        with open(ZIP_LOOKUP_CSV, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                pre = _s(row.get("prefix")).zfill(3)[:3]
                state = _s(row.get("state"))
                if pre and state:
                    table[pre] = (_s(row.get("city")), state)
    except OSError:
        pass  # no bundled file -> prefix->state fallback only
    _ZIP3_CITY_STATE = table
    return table


def _state_from_prefix(prefix: str) -> str:
    try:
        n = int(prefix)
    except ValueError:
        return ""
    for lo, hi, st in _STATE_PREFIX_RANGES:
        if lo <= n <= hi:
            return st
    return ""


def zip_location(zipcode) -> str:
    """Return 'City, ST' (or just 'ST') for a US ZIP, or '' if unknown.

    Offline only and tolerant of messy input (uses the first digits found).
    """
    digits = re.sub(r"\D", "", _s(zipcode))
    if len(digits) < 3:
        return ""
    prefix = digits[:3]
    city, state = _load_zip_lookup().get(prefix, ("", ""))
    if not state:
        state = _state_from_prefix(prefix)
    if city and state:
        return f"{city}, {state}"
    return state or ""


# --------------------------------------------------------------------------- #
# Find the most recent survey-answer export (NOT the consent export)
# --------------------------------------------------------------------------- #
_FNAME_DATE_RE = re.compile(r"_([A-Z][a-z]+ \d{1,2}, \d{4})_(\d{2})\.(\d{2})")


def _export_timestamp(path: str) -> dt.datetime:
    """Parse 'June 25, 2026_06.18' from the filename; fall back to mtime."""
    m = _FNAME_DATE_RE.search(os.path.basename(path).replace("+", " "))
    if m:
        try:
            d = dt.datetime.strptime(m.group(1), "%B %d, %Y")
            return d.replace(hour=int(m.group(2)), minute=int(m.group(3)))
        except ValueError:
            pass
    return dt.datetime.fromtimestamp(os.path.getmtime(path))


def _find_latest_survey_zip():
    candidates = []
    for p in glob.glob(os.path.join(INGEST, "*.zip")):
        base = os.path.basename(p).replace("+", " ")
        if base.lower().startswith("consent"):
            continue                       # consent survey, not the answers
        if "k12" not in base.lower():
            continue
        candidates.append(p)
    if not candidates:
        return None
    return max(candidates, key=_export_timestamp)


def _load_survey_answers():
    """
    Returns (codes, labels, by_cid, export_name) for the latest survey export.
      codes       : column codes        (header row 1)
      labels      : question texts      (header row 2)
      by_cid      : {cid: row-values}   one entry per response
      export_name : basename of the zip used
    """
    path = _find_latest_survey_zip()
    if not path:
        return [], [], {}, None
    with zipfile.ZipFile(path) as zf:
        name = [n for n in zf.namelist() if n.lower().endswith(".csv")][0]
        text = zf.read(name).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    if len(rows) < 4:
        return [], [], {}, os.path.basename(path)
    codes = rows[0]
    labels = rows[1]
    try:
        cid_idx = codes.index("cid")
    except ValueError:
        cid_idx = None
    by_cid = {}
    for r in rows[3:]:
        if cid_idx is not None and cid_idx < len(r):
            cid = _s(r[cid_idx])
            if cid:
                by_cid[cid] = r
    return codes, labels, by_cid, os.path.basename(path)


# --------------------------------------------------------------------------- #
# The in-memory store
# --------------------------------------------------------------------------- #
class Store:
    def __init__(self):
        # tracker
        _, self.tracker = _read_xlsx(TRACKER_XLSX, "Participants")
        self.tracker_by_rid = {_s(r.get("response_id")): r for r in self.tracker}

        # payment tracker
        _, pay = _read_xlsx(PAYMENT_XLSX, "Payments")
        self.pay_by_cid = {_s(r.get("cid")): r for r in pay}

        # fraud blacklist (indexed three ways)
        self.bl_rows = []
        self.bl_by_rid, self.bl_by_email, self.bl_by_ip = {}, {}, {}
        if os.path.exists(BLACKLIST_CSV):
            with open(BLACKLIST_CSV, newline="", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    self.bl_rows.append(row)
                    if _s(row.get("response_id")):
                        self.bl_by_rid[_s(row["response_id"])] = row
                    if _s(row.get("delivery_email")):
                        self.bl_by_email[_norm(row["delivery_email"])] = row
                    if _s(row.get("ip")):
                        self.bl_by_ip[_s(row["ip"])] = row

        # survey answers (latest export)
        self.codes, self.labels, self.ans_by_cid, self.export_name = _load_survey_answers()

    # ------------------------------------------------------------------ #
    # SEARCH
    # ------------------------------------------------------------------ #
    def search(self, query: str) -> dict:
        q = _s(query)
        if not q:
            return {"query": query, "kind": None, "candidates": []}

        # RID?
        if RID_RE.match(q):
            row = self.tracker_by_rid.get(q)
            cands = [self._candidate(row, "response_id")] if row else []
            return {"query": q, "kind": "rid", "candidates": cands}

        # email?
        if "@" in q:
            ql = _norm(q)
            cands = [
                self._candidate(r, "delivery_email")
                for r in self.tracker
                if ql in _norm(r.get("delivery_email"))
            ]
            return {"query": q, "kind": "email", "candidates": self._dedupe(cands)}

        # name (first only, or first+last): every query token must match the
        # START of some word in the name, so "Har" -> "Harper" but "eric"
        # does NOT match "Merica".
        tokens = _norm(q).split()

        def name_match(name) -> bool:
            words = _norm(name).split()
            if not words:
                return False
            return all(any(w.startswith(t) for w in words) for t in tokens)

        cands = []
        for r in self.tracker:
            in_parent = name_match(r.get("parent_name"))
            in_child = name_match(r.get("child_name"))
            if in_parent and in_child:
                matched = "parent + child"
            elif in_parent:
                matched = "parent"
            elif in_child:
                matched = "child"
            else:
                continue
            cands.append(self._candidate(r, matched))
        return {"query": q, "kind": "name", "candidates": self._dedupe(cands)}

    def _dedupe(self, cands):
        seen, out = set(), []
        for c in cands:
            rid = c["response_id"]
            if rid in seen:
                continue
            seen.add(rid)
            out.append(c)
        # completed first, then most recent
        out.sort(key=lambda c: (not c["completed"], c.get("recorded_date") or ""))
        return out

    def _candidate(self, r, matched_on):
        rid = _s(r.get("response_id"))
        status, _ = self._status(rid, r)
        return {
            "response_id": rid,
            "parent_name": _s(r.get("parent_name")),
            "child_name": _s(r.get("child_name")),
            "delivery_email": _s(r.get("delivery_email")),
            "grade": _s(r.get("grade")),
            "matched_on": matched_on,
            "completed": self._is_completed(rid, r),
            "status": status,
            "recorded_date": _s(r.get("recorded_date")),
        }

    # ------------------------------------------------------------------ #
    # STATUS  (precedence: fraud > paid > hold > pay)
    # ------------------------------------------------------------------ #
    def _blacklist_hit(self, rid, r):
        pay = self.pay_by_cid.get(rid, {})
        email = _norm(r.get("delivery_email")) or _norm(pay.get("delivery_email"))
        ip = _s(pay.get("survey_ip"))
        for key, idx in ((rid, self.bl_by_rid), (email, self.bl_by_email), (ip, self.bl_by_ip)):
            if key and key in idx:
                return idx[key]
        return None

    def _status(self, rid, r):
        """Return (STATUS, reason)."""
        pay = self.pay_by_cid.get(rid)
        bl = self._blacklist_hit(rid, r)

        # 1. FRAUD
        if (pay and _yes(pay.get("fraud"))) or bl:
            reasons = []
            if pay and _s(pay.get("fraud_reason")):
                reasons.append(_s(pay["fraud_reason"]))
            if bl and _s(bl.get("reasons")):
                src = []
                if _s(bl.get("response_id")) == rid:
                    src.append("RID")
                if _norm(bl.get("delivery_email")) == _norm(r.get("delivery_email")):
                    src.append("email")
                if _s(bl.get("ip")) and _s(bl.get("ip")) == _s(pay.get("survey_ip") if pay else ""):
                    src.append("IP")
                src_txt = "/".join(src) or "match"
                reasons.append(f"on fraud blacklist ({src_txt}): {_s(bl['reasons'])}")
            return "FRAUD", "; ".join(dict.fromkeys(reasons)) or "flagged as fraud"

        # 2. PAID
        if pay and _yes(pay.get("paid")):
            extra = []
            if _s(pay.get("paid_date")):
                extra.append(f"on {_s(pay['paid_date'])}")
            if _s(pay.get("paid_amount")):
                extra.append(f"amount {_s(pay['paid_amount'])}")
            return "PAID", ("paid " + " ".join(extra)).strip()

        # 3. HOLD
        if pay and _yes(pay.get("exclude_recommended")):
            return "HOLD", _s(pay.get("flag_reasons")) or _s(pay.get("quality_status")) or "exclusion recommended"

        # 4. PAY (only meaningful once a payment-tracker row exists)
        if pay:
            note = _s(pay.get("flag_reasons"))
            return "PAY", (f"cleared; minor flags: {note}" if note else "cleared, no exclusion flags")

        # No payment-tracker row yet -> fall back to the tracker pipeline status
        tstatus = _s(r.get("status")).upper()
        treason = _s(r.get("reason"))
        if tstatus == "SUSPICIOUS":
            return "HOLD", f"tracker: SUSPICIOUS ({treason})" if treason else "tracker: SUSPICIOUS"
        if tstatus in ("INCOMPLETE", "INELIGIBLE"):
            return "NOT EVALUATED", f"tracker: {tstatus} ({treason})" if treason else f"tracker: {tstatus}"
        return "NOT EVALUATED", "no payment-tracker record yet"

    def _is_completed(self, rid, r) -> bool:
        pay = self.pay_by_cid.get(rid)
        if pay and _yes(pay.get("completed")):
            return True
        if _s(r.get("survey_completed_at")):
            return True
        return rid in self.ans_by_cid

    # ------------------------------------------------------------------ #
    # DETAIL
    # ------------------------------------------------------------------ #
    def detail(self, response_id):
        rid = _s(response_id)
        r = self.tracker_by_rid.get(rid)
        if not r:
            return None

        pay = self.pay_by_cid.get(rid, {})
        status, reason = self._status(rid, r)
        completed = self._is_completed(rid, r)

        person = {
            "response_id": rid,
            "parent_name": _s(r.get("parent_name")),
            "child_name": _s(r.get("child_name")),
            "delivery_email": _s(r.get("delivery_email")),
            "grade": _s(r.get("grade")),
            "recorded_date": _s(r.get("recorded_date")),
            "survey_completed_at": _s(r.get("survey_completed_at")) or _s(pay.get("survey_completed_at")),
            "tracker_status": _s(r.get("status")),
            "tracker_reason": _s(r.get("reason")),
        }

        return {
            "person": person,
            "status": status,
            "reason": reason,
            "completed": completed,
            "metadata": self._metadata(rid, r, pay),
            "featured_text": self._featured_text(rid) if completed else [],
            "answers": self._answers(rid) if completed else [],
            "export_name": self.export_name,
        }

    def _featured_text(self, rid):
        """Q62/Q63 long free-text answers, for the card above survey answers."""
        row = self.ans_by_cid.get(rid)
        if not row or not self.codes:
            return []
        idx = {c: i for i, c in enumerate(self.codes)}
        out = []
        for code in FEATURED_TEXT_CODES:
            i = idx.get(code)
            if i is None:
                continue
            value = _s(row[i]) if i < len(row) else ""
            if not value:
                continue
            label = _s(self.labels[i]) if i < len(self.labels) else ""
            out.append({"code": code, "label": label or code, "value": value})
        return out

    def _metadata(self, rid, r, pay):
        out = []

        def add(label, value):
            v = _s(value)
            if v:
                out.append({"label": label, "value": v})

        # payment-tracker quality signals
        add("Quality status", pay.get("quality_status"))
        add("Flags (count)", pay.get("n_flags"))
        add("Flag reasons", pay.get("flag_reasons"))
        add("Exclude recommended", pay.get("exclude_recommended"))
        add("Throwaway email", pay.get("throwaway_email"))
        add("Consent OK", pay.get("consent_ok"))
        add("Survey IP", pay.get("survey_ip"))

        # signals straight from the survey export
        row = self.ans_by_cid.get(rid)
        if row and self.codes:
            idx = {c: i for i, c in enumerate(self.codes)}

            def col(code):
                i = idx.get(code)
                return row[i] if i is not None and i < len(row) else ""

            add("Response type", col("Status"))
            add("Finished", col("Finished"))
            add("Progress (%)", col("Progress"))
            # Self-reported ZIP (Q66), annotated with City, ST via offline lookup.
            zip_raw = _s(col("Q66"))
            if zip_raw:
                loc = zip_location(zip_raw)
                add("Zip code (self-reported)", f"{zip_raw}  ·  {loc}" if loc else zip_raw)
            # Duration shown in minutes rather than raw seconds.
            dur_raw = _s(col("Duration (in seconds)"))
            if dur_raw:
                try:
                    add("Duration (min)", f"{float(dur_raw) / 60:.1f}")
                except ValueError:
                    add("Duration (min)", dur_raw)
            add("reCAPTCHA score", col("Q_RecaptchaScore"))
            add("Duplicate respondent", col("Q_DuplicateRespondent"))
            add("Straightlining (%)", col("Q_StraightliningPercentage"))
            add("Unanswered (%)", col("Q_UnansweredPercentage"))
        return out

    def _answers(self, rid):
        row = self.ans_by_cid.get(rid)
        if not row or not self.codes:
            return []
        out = []
        for i, code in enumerate(self.codes):
            if i < 19 or i > 142:           # question block sits between metadata and embedded data
                continue
            if any(bit in code for bit in TIMING_BITS):
                continue
            if code.startswith("Q_"):       # system quality fields
                continue
            if code in FEATURED_TEXT_CODES:  # shown in their own card above
                continue
            value = _s(row[i]) if i < len(row) else ""
            if not value:
                continue
            label = _s(self.labels[i]) if i < len(self.labels) else ""
            is_other = is_open_ended(code)
            if is_other and label:
                label = label.replace(" - Text", "").strip()
            # Annotate the ZIP-code answer with its city/state (offline lookup).
            if "zip code" in label.lower():
                loc = zip_location(value)
                if loc:
                    value = f"{value}  ·  {loc}"
            out.append({
                "code": code,
                "label": label or code,
                "value": value,
                "other": is_other,
            })
        return out


_STORE = None


def load_all() -> Store:
    global _STORE
    _STORE = Store()
    return _STORE


def get_store() -> Store:
    return _STORE or load_all()


# =========================================================================== #
# OPEN-ENDED (FREE-TEXT) QUESTIONS
# --------------------------------------------------------------------------- #
# The survey's open-ended / free-text questions, gathered in one place so they
# are easy to view and maintain. These are the "Other (please specify)" write-
# ins and any other typed-response fields. Any column whose code ends in
# "_TEXT" is treated as open-ended automatically; the explicit list below is
# the human-readable record of which questions those are, and lets us flag an
# open-ended question even if Qualtrics renames the column.
#
# Codes are from the K12 Privacy and AI Extension export.
# =========================================================================== #
OPEN_ENDED_CODES = {
    "Q5_10_TEXT":   "Which AI tools have you used for schoolwork? - Other",
    "Q7_10_TEXT":   "What AI tool has your school provided/assigned? - Other",
    "Q9_10_TEXT":   "How do you use AI tools for school? - Other",
    "Q10_10_TEXT":  "Why do you use AI tools for school? - Other",
    "Q58_11_TEXT":  "Have you ever changed how you use AI tools? - Other",
    "Q59_11_TEXT":  "Main reason for changing how you use AI tools? - Other",
    "Q21_45_TEXT":  "What kinds of information do AI tools collect? - Other",
    "Q22_15_TEXT":  "Why do AI tools collect information about students? - Other",
}


def is_open_ended(code: str) -> bool:
    """True if a question column is an open-ended / free-text response.

    Matches the explicit OPEN_ENDED_CODES above, plus any Qualtrics "_TEXT"
    write-in column, so newly added free-text fields are handled automatically.
    """
    code = _s(code)
    return code in OPEN_ENDED_CODES or code.endswith("_TEXT")
