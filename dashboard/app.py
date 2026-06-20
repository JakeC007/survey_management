"""
Survey pipeline console — local launcher.

A tiny web app (Python stdlib + openpyxl, both already in the shared venv). It
reads the canonical participant tracker for live numbers and gives you one
button per step that runs the EXISTING scripts in consent_management /
response_management / payment_management. It reimplements no pipeline logic.

Run with run_dashboard.bat (double-click), or:
    ..\\.venv\\Scripts\\python.exe app.py

Then open http://127.0.0.1:5000 (opens automatically). The send/draft actions
drive classic Outlook on Windows, exactly like the .bat files do today.
"""

import json
import os
import subprocess
import sys
import threading
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

try:
    import yaml
except Exception:
    yaml = None

try:
    import openpyxl
except Exception:
    openpyxl = None

HERE = Path(__file__).resolve().parent          # survey_management/dashboard
BASE = HERE.parent                               # survey_management
VENV_PY = BASE / ".venv" / "Scripts" / "python.exe"
TEMPLATE = HERE / "index.html"

PAID_TRUE_VALUES = {"yes", "y", "true", "1", "paid", "done", "sent", "complete"}
FLAGGED_STATUSES = {"INELIGIBLE", "INCOMPLETE"}
# Consent rows the fraud screen marked as bots/scammers (skipped, never emailed).
FRAUD_STATUS = "SUSPICIOUS"

# "Both parties consented" = a signed parent consent AND a signed teen assent.
# The tracker doesn't store the raw consent/assent answers, only status + reason,
# so we infer it. The consent screen (consent_management/send_survey_emails.py)
# writes these fragments into `reason` when a party did NOT properly consent;
# a row with none of them cleared both the consent and the assent gate.
CONSENT_FAIL_MARKERS = ("consent=", "assent=", "parent signature", "child signature")
# Statuses where consent/assent was never validly captured at all (unfinished
# form, or a fraud/bot entry). These never count as consent received.
NO_CONSENT_STATUSES = {"INCOMPLETE", "SUSPICIOUS"}

# Serialize script runs and stop the dashboard from reading the xlsx files while
# a script is writing them. RUN_LOCK rejects a second run; RUNNING makes the
# stats reader return the last snapshot instead of opening a file mid-write.
RUN_LOCK = threading.Lock()
RUNNING = False
_LAST_STATS = None


def load_config():
    cfg = {}
    cfg_path = HERE / "config.yaml"
    if not cfg_path.exists():
        sample = HERE / "config.sample.yaml"
        if sample.exists():
            print("[warn] config.yaml not found — using config.sample.yaml. "
                  "Copy it to config.yaml and add your real survey links.")
            cfg_path = sample
    if yaml and cfg_path.exists():
        try:
            cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        except Exception as e:
            print(f"[warn] could not parse {cfg_path.name}: {e}")
    paths = cfg.get("paths", {})
    server = cfg.get("server", {})
    return {
        "qualtrics": cfg.get("qualtrics_data_sources", {}),
        "tracker": BASE / paths.get("tracker_file", "data/participant_tracker_auto.xlsx"),
        "payment_tracker": BASE / paths.get("payment_tracker_file", "data/payment_tracker.xlsx"),
        "unpaid_report": BASE / paths.get("unpaid_report_file", "data/payment_report_unpaid.xlsx"),
        "ingest": BASE / paths.get("ingest_dir", "ingest"),
        "host": server.get("host", "127.0.0.1"),
        "port": int(server.get("port", 5000)),
        "open_browser": bool(server.get("open_browser", True)),
    }


CFG = load_config()


def _sheet_rows(path, sheet=None):
    if not openpyxl or not Path(path).exists():
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows
    except Exception as e:
        print(f"[warn] could not read {path}: {e}")
        return None


def _idx(rows):
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    return {name: i for i, name in enumerate(hdr)}


def _nonblank(rows, idx, col):
    i = idx.get(col)
    if i is None:
        return 0
    return sum(1 for r in rows[1:] if i < len(r) and r[i] is not None and str(r[i]).strip())


def load_stats():
    global _LAST_STATS
    if RUNNING and _LAST_STATS is not None:
        busy = dict(_LAST_STATS)
        busy["running"] = True
        return busy
    s = {k: 0 for k in ("received_consent", "survey_invited", "survey_completed",
                         "cards_paid", "flagged_ineligible", "flagged_fraud",
                         "pay", "hold", "fraud", "ingest_zips")}
    s["tracker_found"] = False
    s["last_updated"] = ""
    s["running"] = False

    rows = _sheet_rows(CFG["tracker"], "Participants")
    if rows and len(rows) > 1:
        s["tracker_found"] = True
        idx = _idx(rows)
        st = idx.get("status")
        rs = idx.get("reason")

        def _both_consented(r):
            status = (str(r[st]).strip().upper()
                      if st is not None and st < len(r) and r[st] is not None else "")
            if status in NO_CONSENT_STATUSES:
                return False
            reason = (str(r[rs]).strip().lower()
                      if rs is not None and rs < len(r) and r[rs] is not None else "")
            return not any(m in reason for m in CONSENT_FAIL_MARKERS)

        # Count only rows where BOTH the parent consent and the teen assent
        # cleared (signatures + affirmations), not every row in the tracker.
        s["received_consent"] = sum(1 for r in rows[1:] if _both_consented(r))
        s["survey_invited"] = _nonblank(rows, idx, "emailed_at")
        s["survey_completed"] = _nonblank(rows, idx, "survey_completed_at")
        if st is not None:
            s["flagged_ineligible"] = sum(
                1 for r in rows[1:]
                if st < len(r) and str(r[st] or "").strip().upper() in FLAGGED_STATUSES)
            # Fraud is tracked separately from ineligibility: these are the
            # consent rows the bot/fraud screen flagged and refused to email.
            s["flagged_fraud"] = sum(
                1 for r in rows[1:]
                if st < len(r) and str(r[st] or "").strip().upper() == FRAUD_STATUS)

    prows = _sheet_rows(CFG["payment_tracker"], "Payments")
    if prows and len(prows) > 1:
        pi = _idx(prows).get("paid")
        if pi is not None:
            s["cards_paid"] = sum(
                1 for r in prows[1:]
                if pi < len(r) and str(r[pi] or "").strip().lower() in PAID_TRUE_VALUES)

    if openpyxl and Path(CFG["unpaid_report"]).exists():
        try:
            wb = openpyxl.load_workbook(CFG["unpaid_report"], read_only=True)
            for name in wb.sheetnames:
                cnt = max(0, wb[name].max_row - 1)
                if name.lower().startswith("pay"):
                    s["pay"] = cnt
                elif name.lower().startswith("hold"):
                    s["hold"] = cnt
                elif name.lower().startswith("fraud"):
                    s["fraud"] = cnt
            wb.close()
        except Exception as e:
            print(f"[warn] could not read unpaid report: {e}")

    try:
        if Path(CFG["ingest"]).exists():
            s["ingest_zips"] = len(list(Path(CFG["ingest"]).glob("*.zip")))
    except Exception:
        pass
    try:
        if Path(CFG["tracker"]).exists():
            import datetime
            mt = datetime.datetime.fromtimestamp(Path(CFG["tracker"]).stat().st_mtime)
            s["last_updated"] = mt.strftime("%b %d, %I:%M%p").lstrip("0")
    except Exception:
        pass
    _LAST_STATS = s
    return s


# Each action maps to an EXISTING script/bat. "py" runs the venv python; "bat" runs cmd.
ACTIONS = {
    "consent_dry":   {"kind": "py",  "cwd": "consent_management",  "args": ["send_survey_emails.py", "--dry-run"]},
    "consent_draft": {"kind": "bat", "cwd": "consent_management",  "bat": "run_draft.bat"},
    "consent_send":  {"kind": "bat", "cwd": "consent_management",  "bat": "run.bat"},
    "response_dry":   {"kind": "bat", "cwd": "response_management", "bat": "run_responses_dryrun.bat"},
    "response_draft": {"kind": "bat", "cwd": "response_management", "bat": "run_responses_draft.bat"},
    "response_send":  {"kind": "bat", "cwd": "response_management", "bat": "run_responses.bat"},
    "payment_dry":   {"kind": "bat", "cwd": "payment_management",  "bat": "run_payments_dryrun.bat"},
    "payment_build": {"kind": "bat", "cwd": "payment_management",  "bat": "run_payments.bat"},
}


def _build_cmd(key):
    """Resolve an action key to (spec, cwd, cmd) or (None, None, error_string)."""
    spec = ACTIONS.get(key)
    if not spec:
        return None, None, f"Unknown action: {key}"
    cwd = BASE / spec["cwd"]
    if not cwd.exists():
        return None, None, f"Folder not found: {cwd}"
    if spec["kind"] == "py":
        py = str(VENV_PY) if VENV_PY.exists() else sys.executable
        # -u so Python flushes each line; lets us stream output live.
        cmd = [py, "-u"] + spec["args"]
    else:
        if not (cwd / spec["bat"]).exists():
            return None, None, f"Script not found: {cwd / spec['bat']}"
        cmd = ["cmd", "/c", spec["bat"]]
    return spec, cwd, cmd


def log(*a):
    """Print to the dashboard's own terminal, flushed immediately."""
    print("[dash]", *a, file=sys.stderr, flush=True)


def stream_action(key):
    """Generator yielding output text chunks as the subprocess produces them."""
    log(f"stream_action: key={key!r}")
    spec, cwd, cmd = _build_cmd(key)
    if spec is None:
        log(f"  build_cmd error: {cmd}")
        yield cmd + "\n"            # cmd holds the error string here
        return
    log(f"  cwd={cwd}")
    log(f"  cmd={cmd}")

    global RUNNING
    if not RUN_LOCK.acquire(blocking=False):
        log("  blocked: another run already in progress")
        yield "Another step is already running. Let it finish before starting another.\n"
        return
    RUNNING = True
    proc = None
    try:
        yield f"$ {' '.join(cmd)}  (in {spec['cwd']})\n\n"
        # PYTHONUNBUFFERED helps any python invoked from inside .bat scripts too.
        # PYTHONIOENCODING forces child stdout to UTF-8 so box-drawing / unicode
        # chars don't blow up on Windows' default cp1252 console codec.
        env = dict(os.environ, PYTHONUNBUFFERED="1", PYTHONIOENCODING="utf-8")
        log("  launching subprocess...")
        proc = subprocess.Popen(
            cmd, cwd=str(cwd), shell=False, env=env,
            stdin=subprocess.DEVNULL,   # so a trailing `pause` in .bat files reads EOF
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, encoding="utf-8", errors="replace",
        )
        log(f"  launched pid={proc.pid}")
        nlines = 0
        for line in iter(proc.stdout.readline, ""):
            nlines += 1
            yield line
        proc.stdout.close()
        rc = proc.wait(timeout=900)
        log(f"  finished pid={proc.pid} exit={rc} lines={nlines}")
        if nlines == 0:
            yield "(script produced no output)\n"
        yield f"\n[exit {rc}]\n"
    except subprocess.TimeoutExpired:
        log("  TIMEOUT - killing")
        if proc:
            proc.kill()
        yield "\n[timed out after 15 minutes; process killed]\n"
    except Exception as e:
        log(f"  EXCEPTION: {e!r}")
        if proc:
            try:
                proc.kill()
            except Exception:
                pass
        yield f"\n[failed to launch: {e}]\n"
    finally:
        RUNNING = False
        RUN_LOCK.release()
        log("  lock released")


def open_report():
    p = CFG["unpaid_report"]
    if not Path(p).exists():
        return False, f"Report not found yet: {p}\nRun 'Build report' first."
    try:
        os.startfile(str(p))  # type: ignore[attr-defined]
        return True, f"Opened {p}"
    except Exception as e:
        return False, f"Could not open report: {e}"


def _card(label, value, desc="", group=""):
    cls = "stat" + (f" {group}" if group else "")
    return (f'<div class="{cls}"><div class="stat-label">{label}</div>'
            f'<div class="stat-value">{value}</div>'
            f'<div class="stat-desc">{desc}</div></div>')


def _qlinks():
    q = CFG["qualtrics"]
    if not q:
        return ""
    items = []
    for v in q.values():
        url = v.get("data_export_url", "#")
        bad = "REPLACE_WITH" in url or "YOURORG" in url
        cls = "ql warn" if bad else "ql"
        tip = "Link not set yet — edit config.yaml" if bad else v.get("feeds", "")
        items.append(f'<a class="{cls}" href="{url}" target="_blank" rel="noopener" title="{tip}">'
                     f'<span class="ql-name">{v.get("name","Qualtrics survey")}</span>'
                     f'<span class="ql-feeds">{v.get("feeds","")}</span></a>')
    return ('<div class="card ql-card"><div class="card-h">'
            '<span class="step-badge gray">Get data</span>'
            '<span class="muted">Qualtrics &rarr; export ZIP &rarr; drop in ingest/</span></div>'
            '<div class="ql-row">' + "".join(items) + "</div></div>")


def render_page():
    s = load_stats()
    banner = ""
    if not s["tracker_found"]:
        banner = ('<div class="banner">No tracker rows yet at '
                  f'<code>{CFG["tracker"].name}</code>. Run step 1 to build it; '
                  "numbers below are zero until then.</div>")
    cards = "".join([
        _card("IRB consent received", s["received_consent"],
              "Parent consent + teen assent both signed", "invite"),
        _card("Survey invites sent", s["survey_invited"],
              "Teens emailed a survey link", "invite"),
        _card("Surveys completed", s["survey_completed"],
              "Finished Qualtrics responses", "done"),
        _card("Gift cards paid", s["cards_paid"],
              "Incentives sent out"),
        _card("Ineligible for payment", s["flagged_ineligible"],
              "Held: ineligible or incomplete"),
        _card("Flagged as fraud", s["flagged_fraud"],
              "On blacklist, never paid"),
    ])
    html = TEMPLATE.read_text(encoding="utf-8")
    repl = {
        "__BANNER__": banner,
        "__CARDS__": cards,
        "__QLINKS__": _qlinks(),
        "__TRACKER_NAME__": CFG["tracker"].name,
        "__INGEST__": str(s["ingest_zips"]),
        "__LAST_UPDATED__": s["last_updated"] or "never",
        "__PAY__": str(s["pay"]),
        "__HOLD__": str(s["hold"]),
        "__FRAUD__": str(s["fraud"]),
    }
    for k, val in repl.items():
        html = html.replace(k, val)
    return html


class Handler(BaseHTTPRequestHandler):
    def _send(self, code, body, ctype="text/html; charset=utf-8"):
        data = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            self._send(200, render_page())
        elif self.path == "/api/stats":
            self._send(200, json.dumps(load_stats()), "application/json")
        else:
            self._send(404, "not found", "text/plain")

    def _stream(self, generator):
        """Stream text chunks to the client as they are produced."""
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("X-Accel-Buffering", "no")
        self.send_header("Connection", "close")
        self.end_headers()
        nbytes = 0
        try:
            for chunk in generator:
                if not chunk:
                    continue
                data = chunk.encode("utf-8")
                self.wfile.write(data)
                self.wfile.flush()
                nbytes += len(data)
        except (BrokenPipeError, ConnectionResetError):
            log("  client disconnected mid-stream")
        log(f"  stream done, {nbytes} bytes sent")

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0) or 0)
        raw = self.rfile.read(length) if length else b"{}"
        try:
            payload = json.loads(raw or b"{}")
        except Exception:
            payload = {}
        log(f"POST {self.path}  payload={payload}")
        if self.path == "/run":
            self._stream(stream_action(payload.get("action", "")))
            return
        if self.path == "/open-report":
            ok, output = open_report()
        else:
            ok, output = False, "not found"
        self._send(200, json.dumps({"ok": ok, "output": output}), "application/json")

    def log_message(self, *args):
        pass


def main():
    host, port = CFG["host"], CFG["port"]
    if openpyxl is None:
        print("[warn] openpyxl not found — stats show zeros. Run setup.bat to build the venv.")
    url = f"http://{host}:{port}"
    print(f"Survey pipeline console running at {url}")
    print("Press Ctrl+C to stop.")
    if CFG["open_browser"]:
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    srv = ThreadingHTTPServer((host, port), Handler)
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
        srv.server_close()


if __name__ == "__main__":
    main()
